"""
Servicio de exportación a Excel.
"""
import io
from typing import List, Optional
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from app.models.lead import Lead, LeadTab


class ExcelExportService:
    """Servicio para exportar leads a Excel."""

    # Colores corporativos (azul Osmofilter)
    HEADER_COLOR = "1E40AF"  # Azul oscuro
    HEADER_FONT_COLOR = "FFFFFF"
    ALT_ROW_COLOR = "EFF6FF"  # Azul muy claro

    def __init__(self):
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export_leads(
        self,
        leads: List[Lead],
        include_notes: bool = True,
        filename: Optional[str] = None
    ) -> bytes:
        """
        Exporta una lista de leads a un archivo Excel.

        Args:
            leads: Lista de objetos Lead
            include_notes: Si incluir las notas de cada lead
            filename: Nombre del archivo (opcional)

        Returns:
            Bytes del archivo Excel
        """
        # Preparar datos
        data = []
        for lead in leads:
            notes_text = ""
            if include_notes and lead.notes:
                notes_text = "\n---\n".join([
                    f"{n.created_at.strftime('%d/%m/%Y %H:%M')}: {n.content}"
                    for n in sorted(lead.notes, key=lambda x: x.created_at)
                ])

            data.append({
                "Nombre": lead.name,
                "URL": lead.url,
                "Dominio": lead.domain,
                "Email": lead.email or "",
                "Teléfono": lead.phone or "",
                "CIF/NIF": lead.cif or "",
                "Estado": lead.status.name if lead.status else "Sin estado",
                "Pestaña": self._tab_to_spanish(lead.tab),
                "Keyword": lead.found_by_keyword.text if lead.found_by_keyword else "",
                "Fecha encontrado": lead.found_at.strftime("%d/%m/%Y") if lead.found_at else "",
                "Notas": notes_text,
                "Descripción": lead.snippet or ""
            })

        # Crear DataFrame
        df = pd.DataFrame(data)

        # Crear Excel con formato
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"

        # Escribir encabezados
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color=self.HEADER_FONT_COLOR)
            cell.fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.thin_border

        # Escribir datos
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                # Alternar color de filas
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color=self.ALT_ROW_COLOR, end_color=self.ALT_ROW_COLOR, fill_type="solid")

        # Ajustar anchos de columna
        column_widths = {
            "A": 40,  # Nombre
            "B": 50,  # URL
            "C": 25,  # Dominio
            "D": 30,  # Email
            "E": 18,  # Teléfono
            "F": 12,  # CIF
            "G": 15,  # Estado
            "H": 15,  # Pestaña
            "I": 25,  # Keyword
            "J": 15,  # Fecha
            "K": 50,  # Notas
            "L": 60   # Descripción
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Congelar primera fila
        ws.freeze_panes = "A2"

        # Guardar
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def export_all_tabs(
        self,
        leads_by_tab: dict,
        country_name: str = "País"
    ) -> bytes:
        """
        Exporta leads organizados por pestañas en diferentes hojas.

        Args:
            leads_by_tab: Dict con {tab_name: [leads]}
            country_name: Nombre del país para el título

        Returns:
            Bytes del archivo Excel
        """
        output = io.BytesIO()
        wb = Workbook()

        # Eliminar hoja por defecto
        wb.remove(wb.active)

        tab_names = {
            LeadTab.NEW: "Leads Nuevos",
            LeadTab.LEADS: "Leads",
            LeadTab.DOUBTS: "Dudas",
            LeadTab.DISCARDED: "Descartados",
            LeadTab.MARKETPLACE: "Marketplaces"
        }

        for tab, leads in leads_by_tab.items():
            if not leads:
                continue

            sheet_name = tab_names.get(tab, str(tab))[:31]  # Excel límite 31 chars
            ws = wb.create_sheet(title=sheet_name)

            # Preparar datos
            data = []
            for lead in leads:
                data.append({
                    "Nombre": lead.name,
                    "Dominio": lead.domain,
                    "Email": lead.email or "",
                    "Teléfono": lead.phone or "",
                    "Estado": lead.status.name if lead.status else "",
                    "Keyword": lead.found_by_keyword.text if lead.found_by_keyword else "",
                    "Fecha": lead.found_at.strftime("%d/%m/%Y") if lead.found_at else ""
                })

            df = pd.DataFrame(data)

            # Escribir encabezados
            for col, header in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color=self.HEADER_FONT_COLOR)
                cell.fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
                cell.border = self.thin_border

            # Escribir datos
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.thin_border
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color=self.ALT_ROW_COLOR, end_color=self.ALT_ROW_COLOR, fill_type="solid")

            # Ajustar anchos
            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

            ws.freeze_panes = "A2"

        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def _tab_to_spanish(self, tab: LeadTab) -> str:
        """Convierte el enum de tab a español."""
        translations = {
            LeadTab.NEW: "Nuevos",
            LeadTab.LEADS: "Leads",
            LeadTab.DOUBTS: "Dudas",
            LeadTab.DISCARDED: "Descartados",
            LeadTab.MARKETPLACE: "Marketplaces"
        }
        return translations.get(tab, str(tab))
